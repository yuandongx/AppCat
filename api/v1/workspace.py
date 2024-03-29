from pathlib import Path

from pymongo import ReplaceOne
from openpyxl import load_workbook

from utils.excel_handle.rd import read
from utils.get_env import look_up
from api.v1.base import Base

"""
财务视图
"""


class Financial(Base):
    urls = ['financial', r'financial/(\w+)']

    async def get(self):
        page_size = self.get_query_argument('pageSize', self.page_size)
        current_page = self.get_query_argument('currentPage', self.current_page)
        data = await self.find(page_params={"size": page_size, "current": current_page})
        self.json(data)

    async def post(self):

        if _id := self.json_args.pop('_id', None):
            oid = self.object_id(_id)
            res = await self.collection.replace_one({'_id': oid}, self.json_args, upsert=True)
            self.json({'code': 0, '_id': _id, 'modified_count': res.modified_count})
        else:
            res = await self.collection.insert_one(self.json_args)
            self.json({"code": 0, "_id": str(res.inserted_id)})

    async def delete(self, param):
        delete_id = param or self.json_args.get('del_ids')
        if delete_id is not None:
            if isinstance(delete_id, list):
                many = self.many_ids(delete_id)
                result = await self.collection.bulk_write(many)
            else:
                result = await self.collection.delete_one({"_id": self.object_id(delete_id)})
            count = result.deleted_count
            self.json({'code': 0, 'deleted_count': count})
        else:
            self.json({'code': -1, 'deleted_count': -1})

    async def patch(self):
        res = await self.collection.update_one(self.json_args)
        self.json({'code': 0, 'matched_count': res.matched_count})


def get_sheet_names(file_name):
    wb = load_workbook(file_name)
    res = {}
    for i, name in enumerate(wb.sheetnames):
        res[f'{i}'] = name
    return res


async def read_and_save(collection, filename, sheet_names):
    result = read(filename, sheet_names)
    updates = []
    filter_keys = ('CompanyName', 'InvoiceDate', 'InvoiceAmount')
    for item in result:
        filters = {key: item[key] for key in filter_keys}
        updates.append(ReplaceOne(filters, item, upsert=True))
    return await collection.bulk_write(updates)


class Upload(Base):
    urls = ['upload']

    async def get(self):
        sheets = self.get_argument('sheets')
        file_name = self.get_argument('file_name')
        split_ = self.get_argument('split')
        if split_ in sheets:
            sheets = sheets.split(split_)
        else:
            sheets = [sheets]
        file_name = self.get_file_name(file_name)
        if file_name.exists():
            filename = file_name.as_posix()
            result = await read_and_save(self.collection, filename, sheets)
            self.json({'insertCount': result.inserted_count})
        else:
            self.json({'insertCount': 0, 'error': f'服务器上不存在`{file_name}`文件。'})

    def post(self):
        excel = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        rtn = {}
        for file in self.request.files['file']:
            if file['content_type'] == excel:
                file_name = self.get_file_name(file['filename'])
                with file_name.open('wb') as f:
                    f.write(file['body'])
                sheet_names = get_sheet_names(file_name)
                rtn[file['filename']] = sheet_names
            else:
                continue
        self.json(rtn)

    @staticmethod
    def get_file_name(name):
        pth = look_up('data_file_save_path')
        if not pth:
            pth = '.'
        path = Path(pth).joinpath('财务报表')
        if not path.exists():
            path.mkdir()
        return path.joinpath(name)
