from pathlib import Path
import asyncio
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.cell import MergedCell
from .config import STYLE_1, STYLE_2


def get_merged_first_cell(ranges, row, col):
    for item in ranges:
        if item[0][0] <= col <= item[0][1] and item[1][0] <= row <= item[1][1]:
            return item[0][0], item[1][0]
    return None, None


def read(file_name, sheet_names=None):
    wb = load_workbook(filename=file_name)
    rows = []
    for sheetname in wb.sheetnames:
        if sheet_names is not None and sheetname not in sheet_names:
            continue
        sheet = wb[sheetname]
        merge_cells = []
        max_col = sheet.max_column
        max_row = sheet.max_row

        for mcell in sheet.merged_cells:
            merge_cells.append(
                ((mcell.min_col, mcell.max_col), (mcell.min_row, mcell.max_row)))

        if max_col < 11:
            style = STYLE_2
        else:
            style = STYLE_1

        start = style['start']
        for i in range(start, max_row):
            row = {}
            
            for key, value in style['headers'].items():
                
                name = value['name']
                cell = sheet[f'{key}{i}']
                if cell.value is None and isinstance(cell, MergedCell):
                    col_index = column_index_from_string(key)
                    c, r = get_merged_first_cell(merge_cells, i, col_index)
                    if c and r:
                        letter = get_column_letter(c)
                        value = sheet[f'{letter}{r}'].value
                    else:
                        value = None
                else:
                    value = cell.value
                row[name] = value
            if not all(row.values()):
                continue
            row['file_name'] = file_name
            row['sheet_name'] = sheetname
            rows.append(row)
    return rows


if __name__ == '__main__':
    pth = Path(__file__).parent.joinpath('test', '操作信息统计表表.xlsx').as_posix()
    res = read(pth)
    
